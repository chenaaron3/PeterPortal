from openpyxl import load_workbook
# import MySQLdb

def parse_departments(filename):
    course_dept = {}
    with open(filename, "r") as file:
        for line in file:
            words = line.rstrip().split(".")
            course_dept[words[-1].strip()] = words[0].strip()
    return course_dept

def get_yr_qtr(year, quarter):
    if quarter == 'Winter' or quarter == 'Spring':
        return quarter + " " + year[0:2] + year[-2:]
    elif quarter == 'Fall' or quarter == 'Summer':
        return quarter + " " + year[:4]
    else:
        raise Exception('Quarter not valid: not Spring, Winter, Summer, or Fall!')

def get_headings(worksheet):
    headings = []
    for header_row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in header_row:
            headings.append(cell.value)
    print("Header:", headings, '\n')
    return headings

def parse_excel(workbook_name, course_dept):

    f = open("rows.txt", "w")
    wb = load_workbook(workbook_name)
    # print(workbook.sheetnames)

    for worksheet in wb:
        # worksheet = wb.worksheets[0] # DEBUGGING: test on first worksheet only

        for row in worksheet.iter_rows(min_row=2):    # DEBUGGING: max_row=2 -> test on 2nd row only
            row_values = [cell.value for cell in row]
            acadTerm = get_yr_qtr(row_values[0], row_values[1])

            try:
                courseID = course_dept[row_values[2]] + str(row_values[3])
            except:
                courseID = row_values[2] + ' ' + str(row_values[3])
                exception_reason = "No course department available for: {}".format(row_values[2])
                raise Exception(exception_reason)

            courseCode = row_values[4]
            sectionNum = row_values[5]
            instructor = row_values[8]
            classType = row_values[7]

            # Grade Distribution Data
            gradeACount = row_values[9]
            gradeBCount = row_values[10]
            gradeCCount = row_values[11]
            gradeDCount = row_values[12]
            gradeFCount = row_values[13]
            gradePCount = row_values[14]
            gradeNPCount = row_values[15]
            gradeWCount = row_values[16]
            gradedGPAAvg = row_values[17]

            # DEBUGGING: get index, headings, cell values
            # headings = get_headings(worksheet)
            # for i, cell in enumerate(row_values):
            #     print(i,headings[i],cell)

            grade_dist = (acadTerm, courseID, courseCode, sectionNum, instructor, classType, gradeACount, gradeBCount, gradeCCount, gradeDCount, gradeFCount, gradePCount, gradeNPCount, gradeWCount, gradedGPAAvg)
            print(grade_dist)

            joined_str = (';'.join(str(x) for x in grade_dist))
            f.write(joined_str + "\n")
    f.close()


# INSERT INTO grade-distribution (AcadTerm, courseID, courseCode, sectionNum, instructor, classType, gradeACount, gradeBCount, gradeCCount, gradeDCount, gradeFCount, gradePCount, gradeNPCount, gradeWCount, gradedGPAAvg)
# VALUES ("Spring 2020", "I&CSCI33", 36620, "A", "PATTIS, R.", "LEC", 125, 72, 31, 16, 33, 1, 0, 1, 2.84);


if __name__ == "__main__":
    course_dept = parse_departments("uci_department.txt")
    parse_excel("combinedData.xlsx", course_dept)
    print("\nDone")
